import openpyxl

def iter_excel_rows_as_json(file_path: str, sheet_name: str = None):
    """
    逐行读取 Excel，每次返回一个 JSON 对象（Python 字典）
    第一行默认作为表头
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 读取表头
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]

    # 从第二行开始逐行返回
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, cell_value in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i+1}"
            row_dict[key] = cell_value
        yield row_dict


# 使用示例
if __name__ == "__main__":
    excel_path = "/Users/liuqianlong/Documents/code/xhscase/web_file_transfer/data/工作簿1.xlsx"  # macOS 路径示例

    for row_json in iter_excel_rows_as_json(excel_path):
        print(row_json)  # 这里 row_json 就是一个 JSON 对象（字典）
